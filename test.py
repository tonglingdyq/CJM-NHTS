import openpyxl, pathlib, os

workbook_path = str(pathlib.Path(os.getcwd()).parent)+'/Results/Bayesian/test.xlsx'
folder_path = os.path.split(workbook_path)[0] #Extract the folder path for the file
pathlib.Path(folder_path).mkdir(parents=True, exist_ok=True) #Create the folder (and parent folder) if not exists yet 
if os.path.exists(workbook_path): #Remove file if it already exists
	os.remove(workbook_path)

# workbook = openpyxl.Workbook()
# ws0 = workbook.active
# ws0.title = 'General Results'
# workbook.save(workbook_path)

# for i in range(3):
# 	workbook = openpyxl.load_workbook(workbook_path)
# 	ws1 = workbook.create_sheet('mY'+str(i))
# 	current_title = [1,2,3,4,5]
# 	for col_i, cell_val in enumerate(current_title, start = 1):
# 		ws1.cell(row = 1, column=col_i).value = cell_val
# 	workbook.save(workbook_path)